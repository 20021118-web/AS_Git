# -*- coding: utf-8 -*-
"""
engine.py — 부품 이미지 매칭 엔진 (Photo_Val.py 의 SIFT 로직 이식)

  · parse_workbook : 엑셀에서 (셀 삽입 이미지 + 서비스 코드) 추출
  · get_sift_features / match_features : SIFT + RANSAC inlier 매칭 (원본과 동일)
  · remove_bg_white : rembg 배경 제거 + 흰 배경 합성 → JPG
"""
import io

import cv2
import numpy as np
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl_image_loader import SheetImageLoader
from PIL import Image

_sift = cv2.SIFT_create()

# ---------------------------------------------------------------- 엑셀 파싱

def parse_workbook(xlsx_bytes: bytes, img_col: str = "D", code_col: str = "E", start_row: int = 4):
    """엑셀 바이트 → [{sheet, rows:[{row, code, image(PIL)}]}]  (Photo_Val.py 와 동일 규칙)"""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    img_idx = column_index_from_string(img_col.upper())
    out = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            loader = SheetImageLoader(ws)
        except Exception:
            continue

        # 코드가 있는 마지막 행 (원본 last_row 로직)
        last_row = ws.max_row
        while last_row >= 1 and not ws[f"{code_col}{last_row}"].value:
            last_row -= 1

        rows = []
        for row in range(start_row, last_row + 1):
            raw = ws[f"{code_col}{row}"].value
            if raw is None:
                continue
            code = str(raw).strip()
            if not code or code == "None":
                continue
            cell = f"{img_col.upper()}{row}"
            try:
                if not loader.image_in(cell):
                    continue
                pil = loader.get(cell)
            except Exception:
                continue
            rows.append({"row": row, "code": code, "image": pil})
        if rows:
            out.append({"sheet": sheet_name, "rows": rows})
    return out


# ---------------------------------------------------------------- SIFT 매칭 (Photo_Val.py 그대로)

def get_sift_features(pil_img: Image.Image):
    try:
        if pil_img.mode == "RGBA":
            white = Image.new("RGB", pil_img.size, (255, 255, 255))
            white.paste(pil_img, mask=pil_img.split()[3])
            pil_img = white
        else:
            pil_img = pil_img.convert("RGB")

        img = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        h, w = gray.shape
        if max(h, w) > 512:
            s = 512.0 / max(h, w)
            gray = cv2.resize(gray, (int(w * s), int(h * s)))

        kp, des = _sift.detectAndCompute(gray, None)
        return kp, des
    except Exception:
        return None, None


def match_features(kp1, des1, kp2, des2) -> int:
    """KNN ratio test(0.7) → good>=10 → RANSAC homography inlier 수 (원본과 동일)"""
    if des1 is None or des2 is None or len(des1) < 2 or len(des2) < 2:
        return 0
    try:
        bf = cv2.BFMatcher()
        matches = bf.knnMatch(des1, des2, k=2)
        good = [m for pair in matches if len(pair) == 2 for m, n in [pair] if m.distance < 0.7 * n.distance]
        if len(good) >= 10:
            src = np.float32([kp1[m.queryIdx].pt for m in good]).reshape(-1, 1, 2)
            dst = np.float32([kp2[m.trainIdx].pt for m in good]).reshape(-1, 1, 2)
            _, mask = cv2.findHomography(src, dst, cv2.RANSAC, 5.0)
            if mask is not None:
                return int(np.sum(mask))
        return 0
    except Exception:
        return 0


# ---------------------------------------------------------------- 배경 제거 → JPG

_rembg_session = None

def remove_bg_white(pil_img: Image.Image, use_bg: bool = True, quality: int = 92) -> bytes:
    """배경 제거(옵션) 후 흰 배경 합성 → JPG 바이트 (원본 파이프라인과 동일, 저장만 jpg)"""
    global _rembg_session
    img = pil_img
    if max(img.size) > 1024:  # rembg 메모리 보호 (원본과 동일)
        img = img.copy()
        img.thumbnail((1024, 1024), Image.Resampling.LANCZOS)

    if use_bg:
        from rembg import remove, new_session
        if _rembg_session is None:
            _rembg_session = new_session()  # 모델 1회 로딩 후 재사용
        img = remove(img, session=_rembg_session)

    white = Image.new("RGB", img.size, (255, 255, 255))
    if img.mode == "RGBA":
        white.paste(img, mask=img.split()[3])
    else:
        white.paste(img)

    buf = io.BytesIO()
    white.save(buf, "JPEG", quality=quality)
    return buf.getvalue()


# ---------------------------------------------------------------- 유틸

def pil_to_jpeg_b64(pil_img: Image.Image, max_px: int = 240) -> str:
    """표시용 썸네일 base64 (프런트 카드용)"""
    import base64
    img = pil_img.convert("RGB")
    img.thumbnail((max_px, max_px), Image.Resampling.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, "JPEG", quality=80)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()

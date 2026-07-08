import streamlit as st
import io
import zipfile
import gc
import openpyxl
import numpy as np
import cv2
from PIL import Image
from openpyxl_image_loader import SheetImageLoader
from rembg import remove

# --- [매칭 난이도 설정] ---
MIN_INLIER_COUNT = 15 

# OpenCV SIFT 알고리즘 초기화 (캐싱하여 성능 최적화)
@st.cache_resource
def get_sift_detector():
    return cv2.SIFT_create()

sift = get_sift_detector()

def get_sift_features(pil_img):
    try:
        if pil_img.mode == 'RGBA':
            white_bg = Image.new("RGB", pil_img.size, (255, 255, 255))
            white_bg.paste(pil_img, mask=pil_img.split()[3])
            pil_img = white_bg
        else:
            pil_img = pil_img.convert("RGB")
            
        img_cv = np.array(pil_img)
        img_cv = cv2.cvtColor(img_cv, cv2.COLOR_RGB2BGR)
        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        
        h, w = gray.shape
        max_dim = max(h, w)
        if max_dim > 512:
            scale = 512.0 / max_dim
            gray = cv2.resize(gray, (int(w * scale), int(h * scale)))
            
        keypoints, descriptors = sift.detectAndCompute(gray, None)
        return keypoints, descriptors
    except Exception:
        return None, None

def match_features(kp1, des1, kp2, des2):
    if des1 is None or des2 is None or len(des1) < 2 or len(des2) < 2:
        return 0
    try:
        bf = cv2.BFMatcher()
        matches = bf.knnMatch(des1, des2, k=2)
        
        good_matches = []
        for match in matches:
            if len(match) == 2:
                m, n = match
                if m.distance < 0.7 * n.distance:
                    good_matches.append(m)
        
        if len(good_matches) >= 10:
            src_pts = np.float32([kp1[m.queryIdx].pt for m in good_matches]).reshape(-1, 1, 2)
            dst_pts = np.float32([kp2[m.trainIdx].pt for m in good_matches]).reshape(-1, 1, 2)
            
            M, mask = cv2.findHomography(src_pts, dst_pts, cv2.RANSAC, 5.0)
            if mask is not None:
                return np.sum(mask)
        return 0
    except Exception:
        return 0

# --- [웹 UI 구성] ---
st.set_page_config(page_title="코웨이 AS 부품 AI 이미지 변환기", page_icon="?", layout="wide")

st.title("? AS 부품 이미지 자동 매칭 및 배경 제거")
st.markdown("엑셀 파일과 비교할 원본 사진을 업로드하면 AI가 자동으로 매칭하여 배경을 제거한 이미지를 제공합니다.")

st.sidebar.header("? 파일 업로드")
uploaded_excels = st.sidebar.file_uploader("엑셀 리스트 파일 업로드 (.xlsx)", type=["xlsx"], accept_multiple_files=True)
uploaded_images = st.sidebar.file_uploader("비교할 원본 사진 업로드", type=["png", "jpg", "jpeg", "webp"], accept_multiple_files=True)

if st.button("실행 및 결과 파일(ZIP) 생성", type="primary"):
    if not uploaded_excels or not uploaded_images:
        st.warning("엑셀 파일과 원본 사진을 모두 업로드해주세요.")
    else:
        with st.spinner("거대한 AI 엔진을 예열 및 원본 사진을 분석 중입니다... 잠시만 기다려주세요!"):
            folder_images_info = []
            
            # 1단계: 원본 사진 분석
            for img_file in uploaded_images:
                try:
                    img_bytes = img_file.read()
                    img = Image.open(io.BytesIO(img_bytes))
                    kp, des = get_sift_features(img)
                    if kp is not None and des is not None:
                        folder_images_info.append({
                            'filename': img_file.name, 
                            'img_bytes': img_bytes, 
                            'kp': kp, 
                            'des': des
                        })
                except Exception as e:
                    st.error(f"사진 분석 오류 ({img_file.name}): {e}")

            st.success(f"분석 완료: **{len(folder_images_info)}**개의 원본 데이터 장전됨.")
            
            # 메모리 내 ZIP 파일 생성 준비
            zip_buffer = io.BytesIO()
            
            # 2단계: 엑셀 매칭 및 배경 제거
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                for idx, excel_file in enumerate(uploaded_excels):
                    status_text.text(f"처리 중: {excel_file.name}")
                    excel_basename = excel_file.name.rsplit('.', 1)[0]
                    
                    try:
                        wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
                        for sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            try:
                                image_loader = SheetImageLoader(ws)
                            except Exception:
                                continue
                            
                            last_row = ws.max_row
                            for row in range(4, last_row + 1):
                                image_cell_coord = f'D{row}'
                                code_cell_coord = f'E{row}'
                                
                                service_code_raw = ws[code_cell_coord].value
                                if not service_code_raw:
                                    continue
                                    
                                service_code = str(service_code_raw).strip()
                                
                                try:
                                    if not image_loader.image_in(image_cell_coord):
                                        continue
                                    
                                    excel_pil_img = image_loader.get(image_cell_coord)
                                    ex_kp, ex_des = get_sift_features(excel_pil_img)
                                    if ex_kp is None or ex_des is None: continue
                                    
                                    best_match_info = None
                                    max_inliers = -1
                                    
                                    for info in folder_images_info:
                                        inliers = match_features(ex_kp, ex_des, info['kp'], info['des'])
                                        if inliers > max_inliers:
                                            max_inliers = inliers
                                            best_match_info = info
                                            
                                    if best_match_info and max_inliers >= MIN_INLIER_COUNT:
                                        # 매칭 성공 처리
                                        original_img = Image.open(io.BytesIO(best_match_info['img_bytes']))
                                        max_size = 1024
                                        if max(original_img.size) > max_size:
                                            original_img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
                                            
                                        bg_removed_img = remove(original_img)
                                        white_bg = Image.new("RGB", bg_removed_img.size, (255, 255, 255))
                                        
                                        if bg_removed_img.mode == 'RGBA':
                                            white_bg.paste(bg_removed_img, mask=bg_removed_img.split()[3])
                                        else:
                                            white_bg.paste(bg_removed_img)
                                        
                                        # 이미지 바이트 변환 후 ZIP에 추가
                                        img_byte_arr = io.BytesIO()
                                        white_bg.save(img_byte_arr, format='PNG')
                                        zip_file.writestr(f"{excel_basename}/{service_code}.png", img_byte_arr.getvalue())
                                        
                                        del bg_removed_img
                                        del white_bg
                                        gc.collect()
                                        
                                except Exception:
                                    pass # 개별 행 오류는 무시하고 진행
                                    
                    except Exception as e:
                        st.error(f"엑셀 처리 오류 ({excel_file.name}): {e}")
                        
                    progress_bar.progress((idx + 1) / len(uploaded_excels))
            
            status_text.text("모든 작업이 완료되었습니다!")
            st.balloons()
            
            # ZIP 다운로드 버튼 생성
            st.download_button(
                label="? 처리된 이미지 다운로드 (ZIP)",
                data=zip_buffer.getvalue(),
                file_name="processed_as_images.zip",
                mime="application/zip",
                type="primary"
            )